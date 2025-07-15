import os
import re
import subprocess
from fuzz_liboqs import TESTPATHS
import openpyxl
import collections
import itertools
import sqlite3


class TexReport:

    def __init__(self):
        self.section = None
        self.sections = {}
        pass


    def guess_crash(self, rowdata):
        if "error" in rowdata:
            line = rowdata["error"]
            if "SEGV" in line:
                return "segfault"
            if "stack-overflow" in line:
                return "stackoverflow"
            if "heap-buffer-overflow" in line:
                return "heapoverflow"
            if "ERROR:" in line and "failed!" in line:
                return "returnedbot"
            print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
            print("@@@@@@@@@@@@@    OTHER    @@@@@@@@@@@@@@@")
            print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
            return "other"
        if "expected" in rowdata:
            if rowdata["expected"] == "unequal":
                return "maul"
            else:
                return "nonmaul"
        import pprint
        pprint.pprint(rowdata)
        raise Exception("Unknown crash cause.")

    def addentry(self, alg_name, rowdata=None, hang=False):
        if hang:
            self.sections[self.section]["hang"].append(alg_name)
            return

        self.sections[self.section][self.guess_crash(rowdata)].append(alg_name)

    def startsection(self, section):
        if section in self.sections:
            return
        self.section = section
        self.sections[section] = {
            "maul": [],
            "nonmaul": [],
            "segfault": [],
            "stackoverflow": [],
            "heapoverflow": [],
            "hang": [],
            "returnedbot": [],
            "other": []
        }

    def savetodisk(self, filename):
        src = ""
        src += "\\begin{tabular}{lc@{\hskip 0em}c@{\hskip 0em}c}\n"
        src += "& \multicolumn{2}{l}{{\hskip 4em}Bit contribution failure} & \multicolumn{1}{c}{Bit exclusion failure} \\\\\n"
        src += "Test &  Malleabilities & Crashes/hangs & Non-malleabilities \\\\ \hline \hline\n"
        src += "& & &  \\\\\n"
        src += "LIBRARY & & & \\\\ \midrule %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n"

        for test in self.sections:
            mauls = len(list(set(self.sections[test]["maul"])))
            nonmauls = len(list(set(self.sections[test]["nonmaul"])))
            errs = []
            segfault = len(list(set(self.sections[test]["segfault"])))
            if segfault:
                errs.append(f"{segfault} (segfault)")
            stackoverflow = len(list(set(self.sections[test]["stackoverflow"])))
            if stackoverflow:
                errs.append(f"{stackoverflow} (stack overflow)")
            heapoverflow = len(list(set(self.sections[test]["heapoverflow"])))
            if heapoverflow:
                errs.append(f"{heapoverflow} (heap overflow)")
            hang = len(list(set(self.sections[test]["hang"])))
            if hang:
                errs.append(f"{hang} (hang)")
            returnedbot = len(list(set(self.sections[test]["returnedbot"])))
            if returnedbot:
                errs.append(f"{returnedbot} (returns $\\bot$)")
            other = len(list(set(self.sections[test]["other"])))
            if other:
                errs.append(f"{other} (other)")
            if len(errs) == 0:
                errstr = " 0 "
            else:
                errstr = " + ".join(errs)
            src += f"{test} & {mauls} & {errstr} & {nonmauls} \\\\\n"

        src += "\end{tabular}\n"

        # save report
        if os.path.exists(f"{filename}.tex"):
            os.remove(f"{filename}.tex")
        with open(f"{filename}.tex", "w") as f:
            f.write(src)


class XLSXReport:

    def __init__(self):
        self.rowdatalist = []
        self.sections = []

    def addentry(self, rowdata, bold=False, italic=False):
        self.rowdatalist.append((rowdata, bold, italic))

    def startsection(self, rowdata, bold=False, italic=False):
        self.addentry(rowdata, bold=bold, italic=italic)
        self.sections.append(len(self.rowdatalist))

    @staticmethod
    def _excel_col(n):
        h = n // 26
        l = n % 26
        return chr(ord('A') + l) if h == 0 else chr(ord('A') + h - 1) + chr(ord('A') + l)

    def savetodisk(self, filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        colors = itertools.cycle([ "EF9A9A", "CE93D8", "B39DDB", "9FA8DA",
                               "81D4FA", "80CBC4", "A5D6A7", "C5E1A5",
                               "FFF59D", "FFE082", "FFAB91", "BCAAA4",
                               "EEEEEE", "B0BEC5" ])

        row = 0
        max_col = max([len(_[0]) for _ in self.rowdatalist] + [0])
        for rowdata, bold, italic in self.rowdatalist:

            # write data to worksheet
            sheet.append(rowdata)
            row += 1

            # style row
            if bold or italic:
                for ws_row in sheet[f"A{row}:{self._excel_col(max_col)}{row}"]:
                    for cell in ws_row:
                        cell.font = openpyxl.styles.Font(name="Calibri", bold=bold, italic=italic)

            # adjust column sizes
            for col in range(sheet.max_column):
                sheet.column_dimensions[self._excel_col(col)].width = 20

        # color-code sections
        for ind in range(len(self.sections)):
            color = next(colors)
            cur = self.sections[ind]
            _next =  len(self.rowdatalist) + 1  if ind == len(self.sections) - 1 else self.sections[ind+1]
            for _row in range(cur, _next):
                # color row using `color`
                for ws_row in sheet[f"A{_row}:{self._excel_col(max(max_col-1, 25))}{_row}"]:
                    for cell in ws_row:
                        cell.fill = openpyxl.styles.PatternFill(
                                        start_color=color,
                                        end_color=color,
                                        fill_type='solid')

        # save report
        if os.path.exists(f"{filename}.xlsx"):
            os.remove(f"{filename}.xlsx")
        workbook.save(filename=f"{filename}.xlsx")


class SQLiteReport:

    def __init__(self):
        self.section = ""
        self.data = []
        self.TESTCOL = "test"
        self.columns = []

    def setcolumns(self, columns):
        for c in columns:
            _c = f"{self.columnsanitise(c)} TEXT"
            if _c not in self.columns:
                self.columns.append(_c)

    def startsection(self, section):
        self.section = section

    def addentry(self, rowdata):
        rowdata[self.TESTCOL] = self.section
        self.setcolumns(list(rowdata.keys()))
        self.data.append(rowdata)

    @staticmethod
    def columnsanitise(c):
        return c.replace("(", "_").replace(")", "_").replace("/", "_").replace(" ", "_").replace("[", "_").replace("]", "").replace("SIG_", "").replace("KEM_", "")

    def savetodisk(self, filename):
        if os.path.exists(f"{filename}.db"):
            os.remove(f"{filename}.db")
        con = sqlite3.connect(f"{filename}.db")
        cur = con.cursor()
        if len(self.columns) == 0:
            print("No crashes detected.")
            return
        create_query = "CREATE TABLE crashes(%s)" % (", ".join(self.columns))
        print(create_query)
        try:
            cur.execute(create_query)
        except Exception as e:
            print("CREATE:", create_query)
            con.close()
            raise(e)
        cnt = 0
        for d in self.data:
            values_string = ""
            column_string = ""
            try:
                for key in d:
                    if len(column_string) > 0:
                        values_string += ", "
                        column_string += ", "
                    column_string += self.columnsanitise(key)
                    values_string += "?" # str(d[key])
            except Exception as e:
                import pdb; pdb.set_trace()
                con.close()
                raise(e)
            insert_query = f"INSERT INTO crashes({column_string}) VALUES({values_string})"
            try:
                cur.execute(insert_query, list(map(str, d.values())))
                print("inserting:", d[self.TESTCOL])
                cnt += 1
            except Exception as e:
                print("INSERT:", insert_query)
                con.close()
                raise(e)
        print("inserted:", cnt)
        con.commit()
        con.close()


def main(mutator, liboqs, report_fn="crash_report"):
    xlsreport = XLSXReport()
    sqlreport = SQLiteReport()
    texreport = TexReport()

    for testpath in TESTPATHS:
        no_entry_in_path = True
        aggr_dir = os.path.join(testpath, f'aggr_{liboqs}_{mutator}')
        if not os.path.isdir(aggr_dir):
            print(f"{aggr_dir} doesn't exist, skipping")
            continue
        for alg_aggr in sorted(os.listdir(aggr_dir)):
            if alg_aggr == "problematic.txt":
                continue
            # print(alg_aggr)
            alg = int(re.findall(r'\d+', alg_aggr)[0])
            alg_dir = os.path.join(aggr_dir, alg_aggr)
            if not os.path.isdir(alg_dir):
                continue
            with open(os.path.join(alg_dir, 'alg.txt')) as f:
                alg_name = f.readline()
            if alg_name == "DEFAULT":
                continue
            print(f"Including {alg_dir} {alg_name} in report")
            hangs_dir = os.path.join(alg_dir, 'fuzzoutputs', 'default', 'hangs')
            for file in os.listdir(hangs_dir):
                filename = os.fsdecode(file)
                if filename == "README.txt":
                    continue
                # print(f"\t{filename}")

                if no_entry_in_path:
                    # xlsreport.addentry(["path", os.path.join(*testpath.split("/")[-3:])], bold=True)
                    xlsreport.startsection(["path", os.path.join(*testpath.split("/")[-3:])], bold=True)
                    sqlreport.startsection(os.path.join(*testpath.split("/")[-3:]))
                    texreport.startsection(os.path.join(*testpath.split("/")[-3:]))

                # xlsreport.addentry([alg_name], bold=True)
                # xlsreport.addentry(["hang"] * 8, italic=True)
                xlsreport.addentry([alg_name] + (["hang"] * 8))
                # sqlreport.addentry([alg_name] + (["hang"] * 8))
                sqlreport.addentry(collections.OrderedDict({"name": alg_name, "error": "hang"}))
                texreport.addentry(alg_name, hang=True)
                no_entry_in_path = False

            crash_dir = os.path.join(alg_dir, 'fuzzoutputs', 'default', 'crashes')
            for file in os.listdir(crash_dir):
                filename = os.fsdecode(file)
                if filename == "README.txt":
                    continue
                # print(f"\t{filename}")
                # run parser on crash file:

                # 1. make sure parser executable exists (make dirs bin/ParseInput.out)
                parse_env = os.environ.copy()
                parse_env["DIRNAME"] = liboqs
                make = subprocess.run(["make", "clean", "dirs", "bin/ParseInput.out"], stdout=subprocess.PIPE, cwd=os.path.join(testpath, str(alg)), env=parse_env)
                # outs, errs = make.communicate(None)
                # print(make.stdout.decode('ascii'))
                # print("compilation complete")

                # 2. run parser bin and collect output
                parser_exe = os.path.join(testpath, str(alg), 'bin', 'ParseInput.out')
                proc = subprocess.run([parser_exe, os.path.join(crash_dir, filename)], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                out = proc.stdout.decode('ascii')
                err = proc.stderr.decode('ascii')

                # 3. profit
                obj = collections.OrderedDict()
                if "ERROR" in err:
                    for line in err.split('\n'):
                        if "ERROR:" in line:
                            obj["name"] = alg_name
                            obj["error"] = line
                            # don't break, "SUMMARY" in line is preferable
                        if "SUMMARY" in line:
                            obj["name"] = alg_name
                            obj["error"] = line
                            break
                else:
                    for line in out.split('\n'):
                        if line == "":
                            continue
                        try:
                            key, val = tuple(line.split(":"))
                        except Exception as e:
                            print(alg_name)
                            print(line)
                            print(e)
                            continue
                        obj[key] = val.strip()

                if no_entry_in_path:
                    # xlsreport.addentry(["path", os.path.join(*testpath.split("/")[-3:])], bold=True)
                    xlsreport.startsection(["path", os.path.join(*testpath.split("/")[-3:])], bold=True)
                    sqlreport.startsection(os.path.join(*testpath.split("/")[-3:]))
                    texreport.startsection(os.path.join(*testpath.split("/")[-3:]))
                    xlsreport.addentry(list(obj.keys()), bold=True)

                xlsreport.addentry(list(obj.values()))
                sqlreport.addentry(obj)
                texreport.addentry(alg_name, rowdata=obj)
                no_entry_in_path = False

    xlsreport.savetodisk(report_fn)
    sqlreport.savetodisk(report_fn)
    texreport.savetodisk(report_fn)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--liboqs', type=str, default="cur_liboqs")
    parser.add_argument('--mutator', type=str, default="python")

    args = parser.parse_args()
    mutator = args.mutator
    liboqs = args.liboqs

    # print (args)

    main(mutator, liboqs, report_fn=f"reports/crash_report_{liboqs}_{mutator}")
